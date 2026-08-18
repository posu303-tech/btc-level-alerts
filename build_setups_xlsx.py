#!/usr/bin/env python3
"""Reproduce the desk-brief setup table (2026-08-18) as an Excel workbook.

Run: python build_setups_xlsx.py [out.xlsx]   (default: setups_2026-08-18.xlsx)
Regenerates the workbook with the current static/live levels.
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "setups_2026-08-18.xlsx"

DATE = "2026-08-18"
LIVE = "live 04:30 UTC (binance)"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
SUB_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def style_body(ws, first, last, ncols):
    for r in range(first, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def widths(ws, w):
    for i, x in enumerate(w, start=1):
        ws.column_dimensions[get_column_letter(i)].width = x


wb = Workbook()

# ---------------- Sheet 1: Setups ----------------
ws = wb.active
ws.title = "Setups"
ws.append(["ID", "Side", "Setup", "Trigger", "Entry", "Stop", "T1", "T2", "Volume condition", "Status"])
style_header(ws, 1, 10)

rows = [
    [
        "L1",
        "Long",
        "POC/Pivot pullback",
        "Price enters POC 63,875..Pivot zone (63,875-63,964, Pivot rotates) and holds >= 30 min",
        "63,875-63,964",
        "63,319 (S1)",
        "64,610 (prior-day high)",
        "65,178 (R1)",
        "-",
        "Waiting - price 64,130 below zone (LIVE)",
    ],
    [
        "L2",
        "Long",
        "4h-close breakout",
        "4h candle closes above prior-day high 64,610",
        "~64,610 (breakout close)",
        "64,276 (session VWAP)",
        "65,178 (R1)",
        "65,500 (VAH)",
        ">= 1,700 BTC (confirm)",
        "Waiting - price 64,130 below level (LIVE)",
    ],
    [
        "S1",
        "Short",
        "VAH/fib rejection",
        "Price at 65,546 + rejection: bearish 4h close or upper wick > 40% of range",
        "~65,500-65,593 (VAH/fib)",
        "65,823 (R2)",
        "64,610 (prior-day high)",
        "63,964 (Pivot)",
        "< 1,000 BTC (cap)",
        "Waiting - no rejection yet (LIVE)",
    ],
    [
        "S2",
        "Short",
        "4h-close breakdown",
        "4h candle closes below S1 63,319",
        "~63,319 (below S1)",
        "63,964 (Pivot)",
        "62,166 (78.6% fib)",
        "62,105 (S2)",
        ">= 1,500 BTC (confirm)",
        "Waiting - price 64,130 above S1 (LIVE)",
    ],
]
for r in rows:
    ws.append(r)
style_body(ws, 2, 5, 10)
widths(ws, [6, 9, 22, 46, 20, 16, 14, 14, 20, 50])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

note = (
    f"Setups from the desk brief dated {DATE}. POCs/fibs/VAH are static; VWAP, session H/L "
    "and prior-day VWAP rotate intraday (entry references resolve against live levels). "
    "All stop/T1/T2 references are level names - see Levels sheet."
)
ws.cell(row=7, column=1, value=note).font = Font(italic=True, size=9, color="555555")
ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=10)

# ---------------- Sheet 2: Levels ----------------
ws2 = wb.create_sheet("Levels")
ws2.append(["Reference", "Value", "Type", "Notes"])
style_header(ws2, 1, 4)

levels = [
    ["Prior day high", "64,610", "static (8/17 close)", "Pivot input - L2 breakout reference"],
    ["Prior day low", "62,751", "static (8/17 close)", "Pivot input"],
    ["Prior day close", "64,532", "static (8/17 close)", "Pivot input"],
    ["PIVOT", "63,964", "static", "Floor pivot (classic formula)"],
    ["R1", "65,178", "static", "L1 T2, L2 T1"],
    ["R2", "65,823", "static", "S1 stop"],
    ["R3", "67,037", "static", ""],
    ["S1", "63,319", "static", "L1 stop, S2 break level"],
    ["S2", "62,105", "static", "S2 T2"],
    ["S3", "61,460", "static", ""],
    ["Session VWAP", "64,276", f"{LIVE} - rotates", "L2 stop; anchored 00:00 UTC"],
    ["Prior-day VWAP", "63,790", "static (8/17 close)", "Previous full day"],
    ["Session high", "64,578", f"{LIVE} - rotates", ""],
    ["Session low", "64,048", f"{LIVE} - rotates", ""],
    ["POC", "63,875", "static (20-session profile)", "L1 entry min"],
    ["61.8% fib", "65,593", "static (desk brief)", "S1 zone max"],
    ["78.6% fib", "62,166", "static (desk brief)", "S2 T1"],
    ["50% fib", "68,000", "static (desk brief)", ""],
    ["VAH", "65,500", "static (20-session profile)", "L2 T2, S1 zone min"],
    ["ATR (4h)", "live", f"{LIVE}", "Average true range, 14 closed candles"],
    ["SMA9 / SMA20 / SMA50", "63,667 / 63,878 / 63,723", f"{LIVE}", "Closed daily closes - price above all three"],
]
for r in levels:
    ws2.append(r)
style_body(ws2, 2, len(levels) + 1, 4)
widths(ws2, [26, 24, 30, 60])
ws2.freeze_panes = "A2"

n2 = (
    f"Static values are from the desk brief ({DATE}); live values are the monitor's "
    "latest run (binance 04:30 UTC) and rotate - pivot R1-R3/S1-S3 skew by < $100 vs "
    "the brief due to exchange data. See README.md for how each level is computed."
)
ws2.cell(row=len(levels) + 3, column=1, value=n2).font = Font(italic=True, size=9, color="555555")
ws2.merge_cells(start_row=len(levels) + 3, start_column=1, end_row=len(levels) + 3, end_column=4)

# ---------------- Sheet 3: Legend ----------------
ws3 = wb.create_sheet("Legend")
ws3.append(["Reference", "Meaning"])
style_header(ws3, 1, 2)
legend = [
    ["PIVOT, R1-R3, S1-S3", "Floor pivots of the last closed daily candle"],
    ["VWAP", "Session VWAP anchored at 00:00 UTC (includes in-progress 4h candle)"],
    ["PRIOR_VWAP", "Previous full day's VWAP (daily volume-weighted)"],
    ["SESSION_HIGH / SESSION_LOW", "Current session high / low"],
    ["POC / VAH", "Volume-profile point of control / value-area high (20 sessions)"],
    ["FIB618 / FIB786 / FIB50", "Fibonacci levels from the desk brief"],
    ["VAH", "Static 65,500 value-area high from the desk brief"],
    ["Volume CONFIRMED", "4h candle volume >= threshold (L2 1,700 BTC, S2 1,500 BTC)"],
    ["Volume cap (S1)", "Entry only while 4h volume < 1,000 BTC"],
    ["L1 hold rule", "Entry requires price to hold in zone >= 30 min"],
    ["S1 rejection", "Bearish 4h close or upper wick > 40% of the candle's range at VAH/fib zone"],
]
for r in legend:
    ws3.append(r)
style_body(ws3, 2, len(legend) + 1, 2)
widths(ws3, [34, 90])
ws3.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {OUT}")